import os
import re
import io
from io import BytesIO
import requests
import zipfile
import openpyxl
import csv
from bs4 import BeautifulSoup
from urllib.parse import urljoin
from datetime import datetime
import pandas as pd
import numpy as np
from functools import reduce

def energy_data():
  # ========================
  # Download RECS Data Files
  # ========================
  def download_recs_files():
      base_url_for_year = 'https://www.eia.gov/consumption/residential/data/'
      response = requests.get(base_url_for_year)
      response.raise_for_status()
      soup = BeautifulSoup(response.text, 'html.parser')
  
      # Find available RECS years
      year_links = soup.find_all('a', href=re.compile(r'^/consumption/residential/data/\d{4}/$'))
      years = [int(re.search(r'\d{4}', link['href']).group()) for link in year_links]
      if not years:
          print("❌ No RECS years found.")
          return
  
      year = max(years)
      base_path = os.path.join('Energy Data', 'RECS', str(year))
      os.makedirs(base_path, exist_ok=True)
      
      print(f"📂 Saving to {base_path}")
  
      # Navigate to RECS microdata page
      base_url = f'https://www.eia.gov/consumption/residential/data/{year}/'
      microdata_url = f'{base_url}index.php?view=microdata'
      soup = BeautifulSoup(requests.get(microdata_url).text, 'html.parser')
  
      links = soup.find_all('a', href=True)
      zip_file = csv_file = codebook_file = None
  
      for link in links:
          href = link['href']
          text = link.get_text(strip=True).lower()
          if 'zip' in text and href.endswith('.zip'):
              zip_file = href
          elif 'csv' in text and href.endswith('.csv'):
              csv_file = href
          elif ('xlsx' in text or 'codebook' in href.lower()) and href.endswith('.xlsx'):
              codebook_file = href
  
      if zip_file:
          # Download and extract zip file
          full_zip_url = f'{base_url}{zip_file}'
          response = requests.get(full_zip_url)
          response.raise_for_status()
  
          if 'application/zip' not in response.headers.get('Content-Type', ''):
              print("❌ Downloaded file is not a ZIP archive.")
              return
  
          with zipfile.ZipFile(BytesIO(response.content)) as z:
              for file_name in z.namelist():
                  if file_name.endswith('.csv') or file_name.endswith('.txt'):
                      extracted_path = os.path.join(base_path, os.path.basename(file_name))
                      with open(extracted_path, 'wb') as f:
                          f.write(z.read(file_name))
                      print(f"✅ Downloaded {file_name}")
          return
  
      if csv_file:
          full_csv_url = f'{base_url}{csv_file}'
          response = requests.get(full_csv_url)
          response.raise_for_status()
  
          # Manual content inspection
          if b"<html" in response.content[:500].lower():
              # If It's HTML, manually search versions
              version = 5
              last_successful_response = None
              last_successful_file = None
  
              while True:
                  file_name = f'recs{year}_public_v{version}.csv'
                  real_csv_url = f'https://www.eia.gov/consumption/residential/data/{year}/csv/{file_name}'
                  test_response = requests.get(real_csv_url)
  
                  if test_response.status_code != 200 or b"<html" in test_response.content[:500].lower():
                      # Version {version} not found or invalid
                      break
                  else:
                      # Save last successful version
                      last_successful_response = test_response
                      last_successful_file = file_name
                      version += 1  # Keep trying next version
  
              if last_successful_response and last_successful_file:
                  csv_path = os.path.join(base_path, last_successful_file)
                  with open(csv_path, 'wb') as f:
                      f.write(last_successful_response.content)
                  print(f"✅ Downloaded {last_successful_file}")
              else:
                  print("❌ Could not find any working RECS public CSV version.")
                  return
  
          else:
              # Normal CSV download
              csv_path = os.path.join(base_path, os.path.basename(csv_file))
              with open(csv_path, 'wb') as f:
                  f.write(response.content)
              print(f"✅ Downloaded {os.path.basename(csv_file)}")
      else:
          print("❌ Could not find RECS microdata zip or CSV file.")
          return
  
      # Download codebook separately if found
      if codebook_file:
          full_codebook_url = f'{base_url}{codebook_file}'
          response = requests.get(full_codebook_url)
          response.raise_for_status()
  
          if 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' in response.headers.get('Content-Type', ''):
              codebook_path = os.path.join(base_path, os.path.basename(codebook_file))
              with open(codebook_path, 'wb') as f:
                  f.write(response.content)
              print(f"✅ Downloaded {os.path.basename(codebook_file)}")
          else:
              print("❌ Codebook URL did not return an XLSX file.")
      else:
          print("❌ No codebook file found.")
  
  # ========================
  # Download SEDS Data Files
  # ========================
  def download_seds_files():
      seds_url = 'https://www.eia.gov/state/seds/seds-data-complete.php?sid=US'
      soup = BeautifulSoup(requests.get(seds_url).text, 'html.parser')
  
      # Extract latest year from heading
      match = re.search(r'1960-(\d{4})', soup.get_text())
      if not match:
          raise ValueError("❌ Could not find the latest SEDS year.")
      latest_year = match.group(1)
      print(f"Latest SEDS year: {latest_year}")
      print(f'📂 Saving to Energy Data/SEDS/{latest_year}')
  
      base_path = os.path.join('Energy Data', 'SEDS', latest_year)
      os.makedirs(base_path, exist_ok=True)
  
      # Download main CSV
      csv_url = next(
          (urljoin('https://www.eia.gov', a['href']) for a in soup.find_all('a', string='CSV') 
           if 'complete' in a['href'].lower() and 'seds' in a['href'].lower()), None)
      if not csv_url:
          raise ValueError("❌ Could not find Complete_SEDS.csv")
      csv_path = os.path.join(base_path, os.path.basename(csv_url))
      with open(csv_path, 'wb') as f:
          f.write(requests.get(csv_url).content)
      print(f"✅ Downloaded SEDS CSV file")
  
      # Download codes/descriptions
      notes_url = 'https://www.eia.gov/state/seds/seds-technical-notes-complete.php?sid=US'
      soup = BeautifulSoup(requests.get(notes_url).text, 'html.parser')
      code_url = next(
          (urljoin('https://www.eia.gov', a['href']) for a in soup.find_all('a', string='CSV') 
           if 'codes' in a['href'].lower() and 'descriptions' in a['href'].lower()), None)
      if not code_url:
          raise ValueError("❌ Could not find codes/descriptions file.")
      code_path = os.path.join(base_path, os.path.basename(code_url))
      with open(code_path, 'wb') as f:
          f.write(requests.get(code_url).content)
      print(f"✅ Downloaded SEDS Codes and Descriptions")
  
  # ==============================
  # Download Total Energy Data Set
  # ==============================
  def download_total_energy_files():
      base_url = 'https://www.eia.gov/totalenergy/data/monthly/index.php'
      soup = BeautifulSoup(requests.get(base_url).text, 'html.parser')
  
      # Extract the release year from page text
      match = re.search(r'Release Date:\s+\w+\s+\d{1,2},\s+(\d{4})', soup.get_text())
      if not match:
          raise ValueError("❌ Could not find release year.")
      year = match.group(1)
      print(f"Latest Total Energy release year: {year}")
      print(f'📂 Saving to Energy Data/ATB/{year}')
  
      base_path = os.path.join('Energy Data', 'Total Energy', year)
      os.makedirs(base_path, exist_ok=True)
  
      # Download ZIP directly to memory
      zip_tag = soup.find('a', string=re.compile(r'Download all tables ZIP', re.IGNORECASE))
      if not zip_tag or not zip_tag.get('href'):
          raise ValueError("❌ ZIP file link not found.")
      zip_url = urljoin(base_url, zip_tag['href'])
  
      try:
          zip_response = requests.get(zip_url)
          zip_response.raise_for_status()
          zip_bytes = io.BytesIO(zip_response.content)
  
          with zipfile.ZipFile(zip_bytes) as zip_ref:
              for member in zip_ref.infolist():
                  if member.filename.endswith('/') or not member.filename.lower().endswith('.xlsx'):
                      continue  # Skip folders and non-Excel files
  
                  # Read Excel file from ZIP directly
                  xlsx_data = zip_ref.read(member)
                  xlsx_buffer = io.BytesIO(xlsx_data)
                  xlsx_name = os.path.splitext(os.path.basename(member.filename))[0]
  
                  try:
                      wb = openpyxl.load_workbook(xlsx_buffer, data_only=True)
                      sheetnames = wb.sheetnames
                  
                      if "Annual Data" in sheetnames:
                          # Only process "Annual Data" sheet
                          sheet = wb["Annual Data"]
                  
                          # Read from the sheet starting from the correct position
                          label = sheet['A7'].value or 'unknown'
                          label_clean = re.sub(r'[^\w\- ]+', '', str(label)).replace(' ', '_').lower()
                  
                          csv_filename = f"{label_clean}.csv"  # ❗ no prefix like "annual_data" anymore
                          # If it starts with "table_", remove "table_xxx_" prefix
                          if csv_filename.startswith('table_'):
                              # Find the first underscore after "table_", then strip up to there
                              parts = csv_filename.split('_', 2)  # Split into at most 3 parts: "table", "xxx", "rest"
                              if len(parts) == 3:
                                  csv_filename = parts[2]  # Keep only "rest"
                          
                          # Prepend "total_energy_" no matter what
                          csv_filename = f"total_energy_{csv_filename}"
                          csv_path = os.path.join(base_path, csv_filename)
                  
                          with open(csv_path, 'w', newline='', encoding='utf-8') as f:
                              writer = csv.writer(f)
                              for row in sheet.iter_rows(values_only=True):
                                  writer.writerow(row)
                  
                          print(f"✅ Downloaded Total Energy Data: {csv_filename}")
                  
                      else:
                          # If no "Annual Data" sheet, process all sheets individually (ignoring monthly)
                          for sheet_name in sheetnames:
                              if "monthly" in sheet_name.lower():
                                  continue  # ❗ skip any monthly data
                  
                              sheet = wb[sheet_name]
                  
                              # Clean the sheet name for the filename
                              label_clean = re.sub(r'[^\w\- ]+', '', sheet_name).replace(' ', '_').lower()
                  
                              csv_filename = f"{label_clean}.csv"  # ❗ clean filename
                              # If it starts with "table_", remove "table_xxx_" prefix
                              if csv_filename.startswith('table_'):
                                  # Find the first underscore after "table_", then strip up to there
                                  parts = csv_filename.split('_', 2)  # Split into at most 3 parts: "table", "xxx", "rest"
                                  if len(parts) == 3:
                                      csv_filename = parts[2]  # Keep only "rest"
                              
                              # Prepend "total_energy_" no matter what
                              csv_filename = f"total_energy_{csv_filename}"
                              csv_path = os.path.join(base_path, csv_filename)
                  
                              with open(csv_path, 'w', newline='', encoding='utf-8') as f:
                                  writer = csv.writer(f)
                                  for row in sheet.iter_rows(values_only=True):
                                      writer.writerow(row)
                  
                              print(f"✅ Downloaded Total Energy Data: {csv_filename}")
                  
                  except Exception as e:
                      print(f"❌ Failed to process {member.filename}: {e}")
                  
      except requests.RequestException as e:
          print(f"❌ Failed to download ZIP archive: {e}")
          return
  
      # Download Glossary PDF
      pdf_tag = soup.find('a', href=re.compile(r'PDF', re.IGNORECASE), attrs={'title': 'Glossary'})
      if not pdf_tag or not pdf_tag.get('href'):
          raise ValueError("❌ Glossary PDF not found.")
      pdf_url = urljoin(base_url, pdf_tag['href'])
      pdf_path = os.path.join(base_path, os.path.basename(pdf_url))
  
      try:
          with open(pdf_path, 'wb') as f:
              f.write(requests.get(pdf_url).content)
          print(f"✅ Downloaded Total Energy Glossary PDF")
      except Exception as e:
          print(f"❌ Failed to download Glossary PDF: {e}")
  
  # ===========================================
  # Download Latest ATB Workbook & Documentation
  # ===========================================
  def download_atb_files():
      base_path_root = os.path.join('Energy Data', 'ATB')
      base_url_template = 'https://atb.nrel.gov/electricity/{year}/data'
      doc_url = 'https://raw.githubusercontent.com/openEDI/documentation/main/ATB.md'
  
      latest_year = None
      for year in range(2025, 2014, -1):
          url = base_url_template.format(year=year)
          try:
              if requests.head(url, allow_redirects=True, timeout=5).status_code == 200:
                  latest_year = str(year)
                  break
          except requests.RequestException:
              continue
      if not latest_year:
          raise ValueError("❌ No valid ATB year found.")
      print(f"Latest ATB year: {latest_year}")
      print(f'📂 Saving to Energy Data/ATB/{latest_year}')
  
      base_path = os.path.join(base_path_root, latest_year)
      os.makedirs(base_path, exist_ok=True)
  
      soup = BeautifulSoup(requests.get(base_url_template.format(year=latest_year)).text, 'html.parser')
      csv_tag = soup.find('a', string=re.compile(rf'Download the {latest_year} ATB Summary CSV Files', re.IGNORECASE))
      if not csv_tag or not csv_tag.get('href'):
          raise ValueError("❌ ATB Summary CSV not found.")
      csv_url = urljoin(base_url_template.format(year=latest_year), csv_tag['href'])
      csv_path = os.path.join(base_path, os.path.basename(csv_url))
      with open(csv_path, 'wb') as f:
          f.write(requests.get(csv_url).content)
      print(f"✅ Downloaded ATB Summary CSV")
  
      doc_path = os.path.join(base_path, 'ATB.md')
      with open(doc_path, 'wb') as f:
          f.write(requests.get(doc_url).content)
      print(f"✅ Downloaded ATB Documentation")
  
  # ============================
  # Download RMI Dataset Files
  # ============================
  def download_rmi_files():
      base_url = 'https://utilitytransitionhub.rmi.org/data-download/'
      base_dir = os.path.join('Energy Data', 'RMI')
      os.makedirs(base_dir, exist_ok=True)
  
      # Files to keep (original base names)
      keep_basenames = [
          'employees',
          'operations_emissions_by_fuel',
          'revenue_by_tech',
          'utility_state_map'
      ]
  
      try:
          soup = BeautifulSoup(requests.get(base_url).text, 'html.parser')
      except requests.RequestException as e:
          print(f"❌ Failed to retrieve RMI page: {e}")
          return
  
      print('📂 Saving to Energy Data/RMI')
  
      containers = soup.find_all('div', class_='container mb-16')[0]
      for container in containers:
          # Extract "Last updated" year
          match = re.search(r'Last updated:\s+\w+\s+\d{1,2},\s+(\d{4})', container.get_text())
          modified_year = match.group(1) if match else 'unknown'
  
          for a in container.find_all('a', href=True):
              file_url = urljoin(base_url, a['href'])
              file_name = os.path.basename(file_url)
  
              # Only process if .csv, .xlsx, or .zip
              if not re.search(r'\.(csv|xlsx|zip)$', file_name, re.IGNORECASE):
                  continue
  
              base, ext = os.path.splitext(file_name)
              base_lower = base.lower()
  
              # Skip if not in keep list
              if not any(target in base_lower for target in keep_basenames):
                  continue
  
              try:
                  response = requests.get(file_url)
                  response.raise_for_status()
  
                  # Handle zip extraction in memory
                  if ext.lower() == '.zip':
                      zip_bytes = io.BytesIO(response.content)
                      try:
                          with zipfile.ZipFile(zip_bytes, 'r') as zip_ref:
                              for member in zip_ref.infolist():
                                  if member.filename.endswith('/'):
                                      continue  # Skip folders
                                  inner_name = os.path.basename(member.filename)
                                  inner_base, inner_ext = os.path.splitext(inner_name)
  
                                  # Check again inside zip
                                  inner_base_lower = inner_base.lower()
                                  if not any(target in inner_base_lower for target in keep_basenames):
                                      continue
  
                                  # Add year if not already present
                                  if not re.search(r'\d{4}$', inner_base):
                                      inner_name = f"{inner_base}_{modified_year}{inner_ext}"
  
                                  file_path = os.path.join(base_dir, inner_name)
                                  with zip_ref.open(member) as source, open(file_path, 'wb') as target:
                                      target.write(source.read())
                                  print(f"✅ Downloaded RMI Data: {inner_name}")
                      except zipfile.BadZipFile:
                          print(f"❌ Invalid ZIP: {file_name}")
                  else:
                      # Non-zip file
                      if not re.search(r'\d{4}$', base):
                          file_name = f"{base}_{modified_year}{ext}"
                      file_path = os.path.join(base_dir, file_name)
                      with open(file_path, 'wb') as f:
                          f.write(response.content)
                      print(f"✅ Downloaded RMI Data: {file_name}")
  
              except requests.RequestException as e:
                  print(f"❌ Failed to download RMI Data: {file_url}: {e}")
  
  def main():
      try:
          print('Downloading Energy Data ...')
          print('--------------------------------------------------------')
          download_recs_files()
          print('--------------------------------------------------------')
          download_seds_files()
          print('--------------------------------------------------------')
          download_total_energy_files()
          print('--------------------------------------------------------')
          download_atb_files()
          print('--------------------------------------------------------')
          download_rmi_files()
      except Exception as e:
          print(f'An error occurred: {e}')
  # Set base folders
  energy_data_folder = 'Energy Data'
  states = [
      'AK', 'AL', 'AR', 'AZ', 'CA', 'CO', 'CT', 'DC', 'DE', 'FL', 'GA',
      'HI', 'IA', 'ID', 'IL', 'IN', 'KS', 'KY', 'LA', 'MA', 'MD', 'ME',
      'MI', 'MN', 'MO', 'MS', 'MT', 'NC', 'ND', 'NE', 'NH', 'NJ', 'NM',
      'NV', 'NY', 'OH', 'OK', 'OR', 'PA', 'RI', 'SC', 'SD', 'TN', 'TX',
      'UT', 'VA', 'VT', 'WA', 'WI', 'WV', 'WY'
  ]
  
  # =========================
  # Load ATB Data
  # =========================
  
  def load_atb():
      atb_folder = os.path.join(energy_data_folder, 'ATB')
      years = [f for f in os.listdir(atb_folder) if f.isdigit()]
      latest_year = max(years)
      atb_path = os.path.join(atb_folder, latest_year)
  
      csv_files = [f for f in os.listdir(atb_path) if f.endswith('.csv')]
      if not csv_files:
          raise ValueError(f"❌ No CSV files found in {atb_path}")
  
      atb_file = os.path.join(atb_path, csv_files[0])
      print(f"✅ Loading ATB file: {atb_file}")
  
      atb_df = pd.read_csv(atb_file, low_memory=False, index_col=0).reset_index(drop=True)
  
      # Confirm year column
      first_col = atb_df.columns[0]
      if 'year' not in first_col.lower():
          raise ValueError(f"❌ First column '{first_col}' does not contain 'year'.")
  
      # Step 1: Remove 'default' column (no longer needed)
      atb_df = atb_df.drop(columns=['default'])
  
      # Step 2: Create final metadata dict per row
      metadata_rows = []
      for _, row in atb_df.iterrows():
          year = row[first_col]
          metadata = row.drop(first_col).to_dict()
  
          metadata_rows.append({
              'year': int(year),
              'state': 'US',
              'metadata': {'ATB': metadata}
          })
  
      # Step 3: Build final DataFrame
      final_atb_df = pd.DataFrame(metadata_rows)
  
      print("✅ Final ATB metadata dataframe ready!")
      return final_atb_df
  
  # =========================
  # Load RECS Data
  # =========================
  
  def load_recs():
      recs_folder = os.path.join(energy_data_folder, 'RECS')
      years = [f for f in os.listdir(recs_folder) if f.isdigit()]
      latest_year = max(years)
      saved_year = int(latest_year)
      recs_path = os.path.join(recs_folder, latest_year)
  
      recs_files = [f for f in os.listdir(recs_path) if f.endswith('.csv') and 'recs' in f.lower()]
      version_files = []
      for f in recs_files:
          match = re.search(r'v(\d+)', f.lower())
          if match:
              version_files.append((int(match.group(1)), f))
      if not version_files:
          raise ValueError(f"❌ No RECS files with version found in {recs_path}")
  
      highest_version_file = max(version_files, key=lambda x: x[0])[1]
      recs_file = os.path.join(recs_path, highest_version_file)
      print(f"✅ Loading RECS file: {recs_file}")
  
      # Load RECS file
      recs_df = pd.read_csv(recs_file, low_memory=False)
  
      if 'state_postal' not in recs_df.columns:
          raise ValueError("❌ 'state_postal' column missing in RECS file")
  
      # Rename for clarity
      recs_df = recs_df.rename(columns={'state_postal': 'state'})
  
      # Step 1: Create metadata dictionaries
      metadata_rows = []
      for _, row in recs_df.iterrows():
          state = row['state']
          metadata = row.drop('state').to_dict()
  
          # Overwrite/add year manually to be safe
          metadata['year'] = saved_year
  
          metadata_rows.append({
              'year': saved_year,
              'state': state,
              'metadata': {'RECS': metadata}
          })
  
      # Step 2: Build final DataFrame
      final_recs_df = pd.DataFrame(metadata_rows)
  
      print("✅ Final RECS metadata dataframe ready!")
      return final_recs_df
  
  # =========================
  # Load RMI Operations Emissions
  # =========================
  
  def load_rmi():
      # --- Setup ---
      rmi_folder = os.path.join(energy_data_folder, 'RMI')
      print("✅ RMI folder found.")
  
      def make_metadata(df, keyname):
          df = df.copy()
          df['year'] = df['year'].astype(int)
          df['metadata'] = df.drop(columns=['year', 'state']).to_dict(orient='records')
          df['metadata'] = df['metadata'].apply(lambda m: {keyname: m})
          return df[['year', 'state', 'metadata']]
  
      def load_rmi_file(pattern, state_col=None, keyname=None, rename_state=False, is_utility_map=False):
          matches = [f for f in os.listdir(rmi_folder) if pattern in f.lower()]
          if not matches:
              print(f"⚠️ No file matching pattern '{pattern}' found.")
              return None
  
          path = os.path.join(rmi_folder, matches[0])
          print(f"✅ Loading RMI file: {path}")
          df = pd.read_csv(path, dtype=str)
  
          # Special case: utility_state_map needs custom renaming
          if is_utility_map:
              df = df.rename(columns={'state': 'state_name', 'state_abbr': 'state'})
          elif state_col:
              if rename_state:
                  df = df.rename(columns={state_col: 'state'})
              else:
                  df['state'] = df[state_col]
          else:
              df['state'] = "NA"
  
          df['year'] = df['year'].astype(int)
  
          # Convert all numeric columns
          numeric_cols = [col for col in df.columns if col not in ['year', 'state']]
          for col in numeric_cols:
              df[col] = pd.to_numeric(df[col], errors='coerce')
  
          return make_metadata(df, keyname)
  
      # --- Load and process operations_emissions_by_fuel ---
      operations_file = [f for f in os.listdir(rmi_folder) if 'operations_emissions_by_fuel' in f.lower()]
      if not operations_file:
          raise ValueError("❌ No RMI operations_emissions_by_fuel file found.")
      operations_path = os.path.join(rmi_folder, operations_file[0])
      print(f"✅ Loading RMI Operations file: {operations_path}")
      operations_df = pd.read_csv(operations_path, dtype=str)
  
      operations_df['year'] = operations_df['year'].astype(int)
      numeric_columns = [col for col in operations_df.columns if col not in ['year', 'state']]
      for col in numeric_columns:
          operations_df[col] = pd.to_numeric(operations_df[col], errors='coerce')
  
      operations_df_final = make_metadata(operations_df, keyname='RMI_emissions')
  
      # --- Load additional RMI datasets ---
      revenues_df = load_rmi_file("revenue_by_tech", keyname="RMI_revenues")
      employees_df = load_rmi_file("employees", keyname="RMI_employees")
      utility_map_df = load_rmi_file("utility_state_map", keyname="RMI_utility_state_map", is_utility_map=True)
  
      # --- Combine all ---
      all_parts = [operations_df_final, revenues_df, employees_df, utility_map_df]
      combined = pd.concat([df for df in all_parts if df is not None], ignore_index=True)
  
      print("✅ Final RMI metadata dataframe ready!")
      return combined
  
  # =========================
  # Load Total Energy Data
  # =========================
  
  def load_total_energy():
      total_folder = os.path.join(energy_data_folder, 'Total Energy/2025')
      files = [f for f in os.listdir(total_folder) if f.endswith('.csv') and f != 'total_energy_contents.csv']
      all_dfs = []
      
      for file_name in files:
          file_path = os.path.join(total_folder, file_name)
      
          try:
              # Quick scan to detect if 'YYYYMM' format appears (in 3rd row)
              preview = pd.read_csv(file_path, nrows=5, header=None, dtype=str)
              header_preview = preview.iloc[2].fillna('').tolist()
      
              first_col = header_preview[0]
      
              if 'yyyymm' in str(first_col).lower():
                  # ✅ State-organized special case
      
                  # Step 1: Read with header at 3rd row (skip first 2 rows)
                  df = pd.read_csv(file_path, skiprows=2, dtype=str)
      
                  # Step 2: Clean header
                  df.columns = [col.replace(',', '').strip() if isinstance(col, str) else col for col in df.columns]
      
                  # Step 3: Drop 'US' column if it exists
                  df = df.drop(columns=['US'], errors='ignore')
      
                  # Step 4: Reformat 'YYYYMM' to 'YYYY'
                  original_year_col = df.columns[0]
                  df[original_year_col] = df[original_year_col].str[:4]
                  df = df.rename(columns={original_year_col: 'year'})
      
                  # Step 5: Melt into long format
                  value_col_name = file_name.replace('.csv', '').lower()
                  df_melted = df.melt(id_vars=['year'], var_name='state', value_name=value_col_name)
      
                  # Step 6: Convert values to numeric
                  df_melted[value_col_name] = pd.to_numeric(df_melted[value_col_name], errors='coerce')
      
                  # Step 7: Group by (year, state) and take sum
                  df_grouped = df_melted.groupby(['year', 'state'], as_index=True).sum()
      
                  # Step 8: Sort (optional polish)
                  df_grouped = df_grouped.sort_index()
      
                  # Add to the list
                  all_dfs.append(df_grouped)
      
                  print(f"✅ Processed file: {file_name}")
      
              else:
                  # ✅ Normal file case
      
                  # Step 1: Read header rows manually
                  df_header1 = pd.read_csv(file_path, skiprows=8, nrows=1, header=None, dtype=str)
                  df_header2 = pd.read_csv(file_path, skiprows=9, nrows=1, header=None, dtype=str)
      
                  # Step 2: Combine headers
                  combined_headers = []
                  for h1, h2 in zip(df_header1.iloc[0], df_header2.iloc[0]):
                      h1_clean = str(h1).strip().replace(',', '').replace(' ', '_').lower()
                      h2_clean = str(h2).strip().replace(',', '').replace(' ', '_').lower()
                      if h2_clean and h2_clean.lower() != 'nan':
                          combined_headers.append(f"{h1_clean}_{h2_clean}")
                      else:
                          combined_headers.append(h1_clean)
      
                  # Step 3: Read full data
                  df = pd.read_csv(file_path, skiprows=10, names=combined_headers, dtype=str)
      
                  # Step 4: Clean column names
                  col_a_original = combined_headers[0]
      
                  # Step 5: Rename first column to 'year'
                  df = df.rename(columns={col_a_original: 'year'})
      
                  # Step 6: Rename other columns properly
                  cleaned_columns = {}
                  for col in df.columns:
                      if col != 'year':
                          cleaned_columns[col] = f"{file_name.replace('.csv', '').lower()}_{col}_{col_a_original}"
                  df = df.rename(columns=cleaned_columns)
      
                  df['state'] = 'US'
                  df = df.set_index(['year', 'state'])
      
                  all_dfs.append(df)
      
                  print(f"✅ Processed file: {file_name}")
      
          except Exception as e:
              print(f"❌ Failed to process {file_name}: {e}")
      
      # Merge all DataFrames efficiently
      if all_dfs:
          final_df = reduce(lambda left, right: left.join(right, how='outer'), all_dfs)
          final_df = final_df.sort_index()
      
          print("✅ Final Total Energy dataframe ready!")
      
      else:
          final_df = pd.DataFrame()
          print("❌ No data was loaded.")
  
      return final_df
  
  # =========================
  # Load SEDS Data
  # =========================
  
  def load_seds():
      seds_folder = os.path.join(energy_data_folder, 'SEDS')
      years = [f for f in os.listdir(seds_folder) if f.isdigit()]
      latest_year = max(years)
      seds_path = os.path.join(seds_folder, latest_year)
  
      csv_files = [f for f in os.listdir(seds_path) if f.endswith('.csv') and 'complete' in f.lower()]
      if not csv_files:
          raise ValueError(f"❌ No 'complete' CSV files found in {seds_path}")
  
      seds_file = os.path.join(seds_path, csv_files[0])
      print(f"✅ Loading SEDS file: {seds_file}")
  
      seds_df = pd.read_csv(seds_file, low_memory=False)
  
      if 'Data_Status' in seds_df.columns:
          seds_df = seds_df.drop(columns=['Data_Status'])
  
      seds_pivoted = seds_df.pivot_table(index=['Year', 'StateCode'], columns='MSN', values='Data').reset_index()
      seds_pivoted.columns.name = None
      seds_pivoted = seds_pivoted.rename(columns={'Year': 'year', 'StateCode': 'state'})
      final_seds_df = seds_pivoted.set_index(['year', 'state'])
  
      print("✅ Final SEDS dataframe ready!")
      return final_seds_df
  
  # =========================
  # Merge Total Energy & SEDS
  # =========================
  
  def merge_seds_and_total(seds_df, total_df):
      # 🛠 Step 1: Standardize both datasets
      def standardize_index(df):
          if isinstance(df.index, pd.MultiIndex):
              df = df.reset_index()
          else:
              df = df.reset_index()
  
          if 'year' not in df.columns or 'state' not in df.columns:
              raise ValueError("❌ DataFrame missing 'year' or 'state' for merging.")
  
          df['year'] = df['year'].astype(int)
          df['state'] = df['state'].astype(str)
  
          return df.set_index(['year', 'state'])
  
      seds_std = standardize_index(seds_df)
      total_std = standardize_index(total_df)
  
      # 🛠 Step 2: Merge SEDS and Total Energy
      merged = seds_std.join(total_std, how='outer')
  
      # 🛠 Step 3: Reset index (now year and state are columns)
      merged = merged.reset_index()
  
      # 🛠 Step 4: Build metadata column
      priority_cols = ['year', 'state']
      data_cols = [col for col in merged.columns if col not in priority_cols]
  
      def row_to_metadata(row):
          metadata = {col: row[col] for col in data_cols if pd.notna(row[col])}
          return {"SEDS_AND_TOTAL_ENERGY": metadata}
  
      merged['metadata'] = merged.apply(row_to_metadata, axis=1)
  
      # 🛠 Step 5: Keep only year, state, metadata
      final_df = merged[['year', 'state', 'metadata']]
  
      print(f"✅ Merged SEDS and Total Energy into metadata dataframe")
      return final_df
    
  def final_concat():
    # 🛠 Step 1: Concatenate all datasets
    combined_df = pd.concat([
        atb_df,
        recs_df,
        rmi_df,
        seds_and_total_df
    ], axis=0, ignore_index=True)
    
    # Step 2: State abbreviation → full state name
    # Reference: https://www.usps.com/send/official-abbreviations.htm
    state_abbr_to_name = {
        'AL': 'Alabama', 'AK': 'Alaska', 'AZ': 'Arizona', 'AR': 'Arkansas',
        'CA': 'California', 'CO': 'Colorado', 'CT': 'Connecticut', 'DE': 'Delaware',
        'FL': 'Florida', 'GA': 'Georgia', 'HI': 'Hawaii', 'ID': 'Idaho',
        'IL': 'Illinois', 'IN': 'Indiana', 'IA': 'Iowa', 'KS': 'Kansas',
        'KY': 'Kentucky', 'LA': 'Louisiana', 'ME': 'Maine', 'MD': 'Maryland',
        'MA': 'Massachusetts', 'MI': 'Michigan', 'MN': 'Minnesota', 'MS': 'Mississippi',
        'MO': 'Missouri', 'MT': 'Montana', 'NE': 'Nebraska', 'NV': 'Nevada',
        'NH': 'New Hampshire', 'NJ': 'New Jersey', 'NM': 'New Mexico', 'NY': 'New York',
        'NC': 'North Carolina', 'ND': 'North Dakota', 'OH': 'Ohio', 'OK': 'Oklahoma',
        'OR': 'Oregon', 'PA': 'Pennsylvania', 'RI': 'Rhode Island', 'SC': 'South Carolina',
        'SD': 'South Dakota', 'TN': 'Tennessee', 'TX': 'Texas', 'UT': 'Utah',
        'VT': 'Vermont', 'VA': 'Virginia', 'WA': 'Washington', 'WV': 'West Virginia',
        'WI': 'Wisconsin', 'WY': 'Wyoming', 'DC': 'District of Columbia'
    }
    
    # Only replace if it matches a valid abbreviation
    combined_df['state'] = combined_df['state'].apply(
        lambda x: state_abbr_to_name.get(x, x)  # If x is a valid abbreviation, replace; else keep as is
    )
    
    # 🧹 Step 3: Optional sort
    combined_df = combined_df.sort_values(['year', 'state']).reset_index(drop=True)
    
    # ✅ Done
    print(f"✅ Final combined dataset shape: {combined_df.shape}")
    return combined_df
    
  main()
  
  atb_df = load_atb()
  recs_df = load_recs()
  rmi_df = load_rmi()
  
  seds_df = load_seds()
  total_df = load_total_energy()
  seds_and_total_df = merge_seds_and_total(seds_df, total_df)

  final_energy_df = final_concat()

  return final_energy_df
