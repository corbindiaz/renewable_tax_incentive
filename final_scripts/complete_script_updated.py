import requests
import pandas as pd
import time
import zipfile
from io import BytesIO
from skimpy import skim
import os
import io
import openpyxl
import csv
from bs4 import BeautifulSoup
from urllib.parse import urljoin
from datetime import datetime
from functools import reduce
import re
from functools import reduce
import numpy as np
import json
import re
import time
from typing import List, Union, Optional, Tuple
import pandas as pd
from concurrent.futures import ThreadPoolExecutor, as_completed
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout

def law_data():
# DOE and NREL Laws and Incentives Data #
csv_url = "https://developer.nrel.gov/api/transportation-incentives-laws/v1.csv?api_key=GAmcMbhWclW5qULHxvWQWtUw52EsehwTPtfu4cz8&expired=false&incentive_type=GNT%2CTAX%2CLOANS%2CRBATE%2CEXEM%2CTOU%2COTHER&law_type=INC%2CPROG%2CLAWREG%2CSTATEINC&regulation_type=REQ%2CDREST%2CREGIS%2CEVFEE%2CFUEL%2CSTD%2CRFS%2CAIRQEMISSIONS%2CCCEINIT%2CUTILITY%2CBUILD%2CRTC%2COTHER&technology=BIOD%2CETH%2CNG%2CLPG%2CHY%2CELEC%2CPHEV%2CHEV%2CNEVS%2CRD%2CAFTMKTCONV%2CEFFEC%2CIR%2CAUTONOMOUS%2COTHER&user_type=FLEET%2CGOV%2CTRIBAL%2CIND%2CSTATION%2CAFP%2CPURCH%2CMAN%2CMUD%2CTRANS%2COTHER"

# Load the CSV directly into a DataFrame
df = pd.read_csv(csv_url)

# Show the first few rows
print(df.head())

## limits:
# some fields are long text or JSON-like strings 

# save the csv file 
response = requests.get(csv_url)
with open("nrel_laws_incentives.csv", "wb") as f:
    f.write(response.content)

print("CSV file downloaded successfully.")
f.close()

nrel_laws_incentives = pd.read_csv("nrel_laws_incentives.csv")

# Section 48C Energy Communities Data #
url = "http://edx.netl.doe.gov/dataset/22944d5d-d063-4890-a995-064bc59b5a78/resource/3d01f2d6-1c1c-498d-8db2-3a51aa3c07f2/download"

# Spoof headers to mimic a browser
headers = {
    "User-Agent": "Mozilla/5.0",
    "Referer": "http://edx.netl.doe.gov/",
}

response = requests.get(url, headers=headers)
response.raise_for_status()

# Unzip and preview
with zipfile.ZipFile(BytesIO(response.content)) as z:
    print("Files in ZIP:", z.namelist())
    for file in z.namelist():
        if file.endswith(".csv"):
            with z.open(file) as f:
                df = pd.read_csv(f)
                print(f"\nPreview of {file}:")
                print(df.head())
                skim(df)

                filename = os.path.basename(file)  # handles nested folders inside zip
                df.to_csv(filename, index=False)
                print(f"Saved {filename}")



# initial check 
# missingness is only present in one variable, there is a lot of missingness so need to see if it is valuable information to our mission 
# good mix of numerical and string variables 

data_48c = pd.read_csv("48C_CensusTractDesignation.csv")

# first changing datasets to have the same state name 
abbr_to_state = {
    "AL": "Alabama", "AK": "Alaska", "AZ": "Arizona", "AR": "Arkansas", "CA": "California",
    "CO": "Colorado", "CT": "Connecticut", "DE": "Delaware", "FL": "Florida", "GA": "Georgia",
    "HI": "Hawaii", "ID": "Idaho", "IL": "Illinois", "IN": "Indiana", "IA": "Iowa",
    "KS": "Kansas", "KY": "Kentucky", "LA": "Louisiana", "ME": "Maine", "MD": "Maryland",
    "MA": "Massachusetts", "MI": "Michigan", "MN": "Minnesota", "MS": "Mississippi", "MO": "Missouri",
    "MT": "Montana", "NE": "Nebraska", "NV": "Nevada", "NH": "New Hampshire", "NJ": "New Jersey",
    "NM": "New Mexico", "NY": "New York", "NC": "North Carolina", "ND": "North Dakota", "OH": "Ohio",
    "OK": "Oklahoma", "OR": "Oregon", "PA": "Pennsylvania", "RI": "Rhode Island", "SC": "South Carolina",
    "SD": "South Dakota", "TN": "Tennessee", "TX": "Texas", "UT": "Utah", "VT": "Vermont",
    "VA": "Virginia", "WA": "Washington", "WV": "West Virginia", "WI": "Wisconsin", "WY": "Wyoming",  "US": "United States of America", "DC": "District of Columbia"
}

def abbreviation_to_state(abbreviation):
    if isinstance(abbreviation, str):
        abbreviation = abbreviation.upper().strip()  # Make sure it's clean
        return abbr_to_state.get(abbreviation, "State not found")
    else:
        return abbreviation  # leave non-strings unchanged


# Directly modify the 'State' column with abbreviations
nrel_laws_incentives['State'] = nrel_laws_incentives['State'].apply(abbreviation_to_state)
nrel_laws_incentives.to_csv("nrel_laws_incentives.csv", index = False)

# getting year and changing name of state values for all of these 
data_48c['date_last_'] = pd.to_datetime(data_48c['date_last_'])
data_48c['year'] = data_48c['date_last_'].dt.year
data_48c = data_48c.rename(columns={'State_Name': 'State'})
data_48c.to_csv("48C_CensusTractDesignation.csv", index = False)

nrel_laws_incentives['year'] = pd.to_datetime(nrel_laws_incentives['Status Date']).dt.year
nrel_laws_incentives['year'] = nrel_laws_incentives['year'].astype('Int64')
nrel_laws_incentives.to_csv("nrel_laws_incentives.csv", index = False)

data_48c.columns = data_48c.columns + '_data48c'
nrel_laws_incentives.columns = nrel_laws_incentives.columns + '_nrel_incentives'
data_48c = data_48c.rename(columns={'State_data48c': 'State', 'year_data48c': 'year'})
nrel_laws_incentives = nrel_laws_incentives.rename(columns={'State_nrel_incentives': 'State', 'year_nrel_incentives': 'year'})


merged = pd.merge(data_48c, nrel_laws_incentives, on = ['State', 'year'], how = "outer", suffixes = ('_data48c', '_nrel_incentives'))
merged.insert(0, 'year', merged.pop('year'))  # Move 'year' to the front
merged.insert(1, 'state', merged.pop('State'))
merged.columns = merged.columns.str.lower()
merged['year'] = merged['year'].astype('Int64')
merged['law id_nrel_incentives'] = merged['law id_nrel_incentives'].astype('Int64')
merged['sequence number_nrel_incentives'] = merged['sequence number_nrel_incentives'].astype('Int64')

merged = merged.drop(columns = ['oid__data48c', 'ctract_geo_data48c', 'state_fip_data48c', 'county_fip_data48c', 'tract_fip_data48c', 'tract_name_data48c', 
                                'f48c_tract_data48c', 'dataset_ve_data48c', 'shape_leng_data48c', 'shape_area_data48c', 'topic_nrel_incentives', 'technology categories_nrel_incentives', 'user categories_nrel_incentives'])

columns_to_exclude = ['state', 'year']

# Select all columns except the ones to exclude and handle NaN values
merged['Combined'] = merged.drop(columns=columns_to_exclude).apply(
    lambda row: '|'.join(row.dropna().astype(str)), axis=1
)

# Show only the excluded columns and the new 'Combined' column
merged_filtered = merged[columns_to_exclude + ['Combined']]
return merged_filtered

def commmunity_data():
# Define the REST API endpoint
url = "https://arcgis.netl.doe.gov/server/rest/services/Hosted/2024_MSAs_NonMSAs_that_only_meet_the_FFE_Threshold/FeatureServer/0/query"

# Set the parameters for the query
params = {
    "where": "1=1",  # Retrieves all records
    "outFields": "*",  # Retrieves all fields
    "f": "json"  # Specifies the format as JSON
}

# Make the GET request to the API
response = requests.get(url, params=params)
data = response.json()

# Extract the features from the response
features = data['features']

# Normalize the JSON data into a pandas DataFrame
df = pd.json_normalize(features)

# Save the DataFrame to a CSV file
df.to_csv("FFE_Threshold_Areas.csv", index=False)

print("✅ Data downloaded and saved as 'FFE_Threshold_Areas.csv'")


######
# Define the REST API endpoint
url = "https://arcgis.netl.doe.gov/server/rest/services/Hosted/2024_MSAs_NonMSAs_that_are_Energy_Communities/FeatureServer/0/query"

# Set the parameters for the query
params = {
    "where": "1=1",  # Retrieves all records
    "outFields": "*",  # Retrieves all fields
    "f": "json"  # Specifies the format as JSON
}

# Make the GET request to the API
response = requests.get(url, params=params)
data = response.json()

# Extract the features from the response
features = data['features']

# Normalize the JSON data into a pandas DataFrame
df = pd.json_normalize(features)

# Save the DataFrame to a CSV file
df.to_csv("Energy_Communities_2024.csv", index=False)

print("✅ Data downloaded and saved as 'Energy_Communities_2024.csv'")


#######
# ArcGIS REST API endpoint
url = "https://arcgis.netl.doe.gov/server/rest/services/Hosted/US_Power_Plants/FeatureServer/0/query"

# Parameters for the request
params = {
    "where": "LOWER(plant_status) = 'closed'",
    "outFields": "*",  # Get all available fields
    "f": "json",
    "resultRecordCount": 8000,
    "orderByFields": "objectid"
}

# Send request
response = requests.get(url, params=params)
data = response.json()

# Extract features safely
features = data.get("features", [])

if not features:
    print("❌ No data found.")
else:
    # Each feature should have an 'attributes' dictionary
    records = [f["attributes"] for f in features if "attributes" in f]
    df = pd.DataFrame(records)

    # Save to CSV
    df.to_csv("Closed_US_Power_Plants.csv", index=False)
    print("✅ Data saved as 'Closed_US_Power_Plants.csv'")


#######
# Define the REST API endpoint
url = "https://arcgis.netl.doe.gov/server/rest/services/Hosted/ManFacNAICS3133_Emissions/FeatureServer/0/query"

# Set the query parameters (JSON format, not PBF)
params = {
    "where": "LOWER(status_48c) = 'eligible for 48c tax credit as a designated energy community'",
    "outFields": "naicc_ss_desc,objectid_12",
    "orderByFields": "objectid_12",
    "f": "json",
    "resultRecordCount": 8000
}

# Send GET request
response = requests.get(url, params=params)
data = response.json()

# Extract features safely
features = data.get("features", [])
if not features:
    print("❌ No matching records found.")
else:
    records = [f["attributes"] for f in features if "attributes" in f]
    df = pd.DataFrame(records)

    # Save to CSV
    df.to_csv("Eligible_48C_Manufacturing_Facilities.csv", index=False)
    print("✅ Saved as 'Eligible_48C_Manufacturing_Facilities.csv'")


######
# REST API endpoint
url = "https://arcgis.netl.doe.gov/server/rest/services/Hosted/US_generators_coal/FeatureServer/0/query"

# Define parameters to return full data
params = {
    "where": "((f_860m_retirementyear <= 2030) AND (f_860m_nameplatecapacity_mw_ >= 53) AND (UPPER(f_860m_operationalstatus_aug_05) IN ('RETIRED','CANCELEDORPOSTPONED','OPERABLE')) AND (solar_mean >= 1947) AND (forest_land <= 18) AND (elev_var_norm <= 0.09))",
    "outFields": "*",
    "returnGeometry": False,
    "f": "json"
}

# Make the GET request
response = requests.get(url, params=params)
data = response.json()

# Extract attributes
features = data.get("features", [])
records = [f["attributes"] for f in features]

# Convert to DataFrame and save
df = pd.DataFrame(records)
df.to_csv("Filtered_US_Coal_Generators.csv", index=False)
print("✅ Saved as 'Filtered_US_Coal_Generators.csv'")


######
# Base URL for the layer
base_url = "https://arcgis.netl.doe.gov/server/rest/services/Hosted/US_generators_coal/FeatureServer/0/query"

# Parameters common to each request
base_params = {
    "f": "json",  # use 'json' for easy parsing
    "where": "1=1",
    "outFields": "*",
    "returnGeometry": False,
    "spatialRel": "esriSpatialRelIntersects",
    "orderByFields": "objectid ASC",  # ensure consistent paging
}

all_records = []
offset = 0
page_size = 1000  # You can adjust this based on server limits

while True:
    print(f"🔄 Fetching records {offset} to {offset + page_size - 1}")

    # Merge base parameters with paging
    params = {**base_params, "resultOffset": offset, "resultRecordCount": page_size}
    response = requests.get(base_url, params=params)
    data = response.json()

    features = data.get("features", [])
    if not features:
        break  # Exit when no more data

    records = [f["attributes"] for f in features]
    all_records.extend(records)

    offset += page_size
    time.sleep(0.2)  # be kind to the server

# Save to CSV
df = pd.DataFrame(all_records)
df.to_csv("All_US_Coal_Generators.csv", index=False)
print(f"✅ Download complete. Total records: {len(df)}")

# List of dataframes
all_dataframes = {
    "closed_us_power_plants": "Closed_US_Power_Plants.csv",
    "eligible_48c_mfg_facilities": "Eligible_48C_Manufacturing_Facilities.csv",
    "filtered_us_coal_generators": "Filtered_US_Coal_Generators.csv",
    "all_us_coal_generators": "All_US_Coal_Generators.csv",
    "ffe_threshold_areas_csv": "FFE_Threshold_Areas.csv",
    "energy_communities_2024_csv": "Energy_Communities_2024.csv",
}

# List to store the reshaped mini-dataframes
final_rows = []

# Function to find a "state" column (flexible)
def find_state_column(df):
    for col in df.columns:
        if 'state' in col.lower():
            return col
    return None

# Function to find a "year" column (flexible)
def find_year_column(df):
    for col in df.columns:
        if 'year' in col.lower():
            return col
    return None

# Process each dataframe
for name, file_path in all_dataframes.items():
    print(f"\n🔍 Processing {name}...")

    # Read the CSV file into a DataFrame
    df = pd.read_csv(file_path)

    # Try to find 'year' and 'state' columns
    state_col = find_state_column(df)
    year_col = find_year_column(df)

    if state_col is None:
        df['state'] = np.nan
    else:
        df['state'] = df[state_col]

    if year_col is None:
        df['year'] = np.nan
    else:
        df['year'] = df[year_col]

    # Drop extracted columns from the data part to avoid duplication
    data_cols = [col for col in df.columns if col not in ['state', 'year']]

    for _, row in df.iterrows():
        row_data = row[data_cols].to_dict()
        final_rows.append({
            'state': row['state'],
            'year': row['year'],
            'data': json.dumps(row_data, default=str)  # Serialize all other columns as JSON
        })

# Create the final mega-DataFrame
final_df = pd.DataFrame(final_rows)

print("\n✅ All datasets have been processed and combined!")

def convert_year_to_int(df, year_col):
    # Convert to numeric, forcing NaNs where conversion fails
    df[year_col] = pd.to_numeric(df[year_col], errors='coerce', downcast='integer')

    # Remove any float-specific `.0` by converting to integer type explicitly
    df[year_col] = df[year_col].astype('Int64')  # Use pandas' nullable Int64 dtype for handling NaNs
    return df

final_df = convert_year_to_int(final_df, "year")

state_lookup = {
    "AL": "Alabama", "AK": "Alaska", "AZ": "Arizona", "AR": "Arkansas",
    "CA": "California", "CO": "Colorado", "CT": "Connecticut", "DE": "Delaware",
    "FL": "Florida", "GA": "Georgia", "HI": "Hawaii", "ID": "Idaho",
    "IL": "Illinois", "IN": "Indiana", "IA": "Iowa", "KS": "Kansas",
    "KY": "Kentucky", "LA": "Louisiana", "ME": "Maine", "MD": "Maryland",
    "MA": "Massachusetts", "MI": "Michigan", "MN": "Minnesota", "MS": "Mississippi",
    "MO": "Missouri", "MT": "Montana", "NE": "Nebraska", "NV": "Nevada",
    "NH": "New Hampshire", "NJ": "New Jersey", "NM": "New Mexico", "NY": "New York",
    "NC": "North Carolina", "ND": "North Dakota", "OH": "Ohio", "OK": "Oklahoma",
    "OR": "Oregon", "PA": "Pennsylvania", "RI": "Rhode Island", "SC": "South Carolina",
    "SD": "South Dakota", "TN": "Tennessee", "TX": "Texas", "UT": "Utah",
    "VT": "Vermont", "VA": "Virginia", "WA": "Washington", "WV": "West Virginia",
    "WI": "Wisconsin", "WY": "Wyoming",
    "DC": "District of Columbia",
}

def find_year(row):
    for value in row.values():
        if isinstance(value, int) and 1900 <= value <= 2100:
            return value
        if isinstance(value, str) and value.isdigit():
            year = int(value)
            if 1900 <= year <= 2100:
                return year
    return None

def find_state(row):
    for value in row.values():
        if isinstance(value, str):
            val = value.strip().upper()
            if val in state_lookup:
                return state_lookup[val]
            if val.title() in state_lookup.values():
                return val.title()
    return "Federal"

def process_dataframe(df):
    records = []
    for _, row in df.iterrows():
        row_dict = row.to_dict()
        year = find_year(row_dict)
        state = find_state(row_dict)
        rest_of_data = row_dict  # could also remove year/state fields if needed
        records.append({
            "year": year,
            "state": state,
            "rest_of_data": rest_of_data
        })
    return pd.DataFrame(records)

# Example usage:
processed_df = process_dataframe(final_df)
return processed_df

def energy_data():
# ========================
# Download RECS Data Files
# ========================
    def download_recs_files():
        base_url_for_year = 'https://www.eia.gov/consumption/residential/data/'
        response = requests.get(base_url_for_year)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, 'html.parser')

        # Find available RECS years
        year_links = soup.find_all('a', href=re.compile(r'^/consumption/residential/data/\d{4}/$'))
        years = [int(re.search(r'\d{4}', link['href']).group()) for link in year_links]
        if not years:
            print("❌ No RECS years found.")
            return

    year = max(years)
    base_path = os.path.join('Energy Data', 'RECS', str(year))
    os.makedirs(base_path, exist_ok=True)
    
    print(f"📂 Saving to {base_path}")

    # Navigate to RECS microdata page
    base_url = f'https://www.eia.gov/consumption/residential/data/{year}/'
    microdata_url = f'{base_url}index.php?view=microdata'
    soup = BeautifulSoup(requests.get(microdata_url).text, 'html.parser')

    links = soup.find_all('a', href=True)
    zip_file = csv_file = codebook_file = None

    for link in links:
        href = link['href']
        text = link.get_text(strip=True).lower()
        if 'zip' in text and href.endswith('.zip'):
            zip_file = href
        elif 'csv' in text and href.endswith('.csv'):
            csv_file = href
        elif ('xlsx' in text or 'codebook' in href.lower()) and href.endswith('.xlsx'):
            codebook_file = href

    if zip_file:
        # Download and extract zip file
        full_zip_url = f'{base_url}{zip_file}'
        response = requests.get(full_zip_url)
        response.raise_for_status()

        if 'application/zip' not in response.headers.get('Content-Type', ''):
            print("❌ Downloaded file is not a ZIP archive.")
            return

        with zipfile.ZipFile(BytesIO(response.content)) as z:
            for file_name in z.namelist():
                if file_name.endswith('.csv') or file_name.endswith('.txt'):
                    extracted_path = os.path.join(base_path, os.path.basename(file_name))
                    with open(extracted_path, 'wb') as f:
                        f.write(z.read(file_name))
                    print(f"✅ Downloaded {file_name}")
        return

    if csv_file:
        full_csv_url = f'{base_url}{csv_file}'
        response = requests.get(full_csv_url)
        response.raise_for_status()

        # Manual content inspection
        if b"<html" in response.content[:500].lower():
            # If It's HTML, manually search versions
            version = 5
            last_successful_response = None
            last_successful_file = None

            while True:
                file_name = f'recs{year}_public_v{version}.csv'
                real_csv_url = f'https://www.eia.gov/consumption/residential/data/{year}/csv/{file_name}'
                test_response = requests.get(real_csv_url)

                if test_response.status_code != 200 or b"<html" in test_response.content[:500].lower():
                    # Version {version} not found or invalid
                    break
                else:
                    # Save last successful version
                    last_successful_response = test_response
                    last_successful_file = file_name
                    version += 1  # Keep trying next version

            if last_successful_response and last_successful_file:
                csv_path = os.path.join(base_path, last_successful_file)
                with open(csv_path, 'wb') as f:
                    f.write(last_successful_response.content)
                print(f"✅ Downloaded {last_successful_file}")
            else:
                print("❌ Could not find any working RECS public CSV version.")
                return

        else:
            # Normal CSV download
            csv_path = os.path.join(base_path, os.path.basename(csv_file))
            with open(csv_path, 'wb') as f:
                f.write(response.content)
            print(f"✅ Downloaded {os.path.basename(csv_file)}")
    else:
        print("❌ Could not find RECS microdata zip or CSV file.")
        return

    # Download codebook separately if found
    if codebook_file:
        full_codebook_url = f'{base_url}{codebook_file}'
        response = requests.get(full_codebook_url)
        response.raise_for_status()

        if 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' in response.headers.get('Content-Type', ''):
            codebook_path = os.path.join(base_path, os.path.basename(codebook_file))
            with open(codebook_path, 'wb') as f:
                f.write(response.content)
            print(f"✅ Downloaded {os.path.basename(codebook_file)}")
        else:
            print("❌ Codebook URL did not return an XLSX file.")
    else:
        print("❌ No codebook file found.")

# ========================
# Download SEDS Data Files
# ========================
def download_seds_files():
    seds_url = 'https://www.eia.gov/state/seds/seds-data-complete.php?sid=US'
    soup = BeautifulSoup(requests.get(seds_url).text, 'html.parser')

    # Extract latest year from heading
    match = re.search(r'1960-(\d{4})', soup.get_text())
    if not match:
        raise ValueError("❌ Could not find the latest SEDS year.")
    latest_year = match.group(1)
    print(f"Latest SEDS year: {latest_year}")
    print(f'📂 Saving to Energy Data/SEDS/{latest_year}')

    base_path = os.path.join('Energy Data', 'SEDS', latest_year)
    os.makedirs(base_path, exist_ok=True)

    # Download main CSV
    csv_url = next(
        (urljoin('https://www.eia.gov', a['href']) for a in soup.find_all('a', string='CSV') 
        if 'complete' in a['href'].lower() and 'seds' in a['href'].lower()), None)
    if not csv_url:
        raise ValueError("❌ Could not find Complete_SEDS.csv")
    csv_path = os.path.join(base_path, os.path.basename(csv_url))
    with open(csv_path, 'wb') as f:
        f.write(requests.get(csv_url).content)
    print(f"✅ Downloaded SEDS CSV file")

    # Download codes/descriptions
    notes_url = 'https://www.eia.gov/state/seds/seds-technical-notes-complete.php?sid=US'
    soup = BeautifulSoup(requests.get(notes_url).text, 'html.parser')
    code_url = next(
        (urljoin('https://www.eia.gov', a['href']) for a in soup.find_all('a', string='CSV') 
        if 'codes' in a['href'].lower() and 'descriptions' in a['href'].lower()), None)
    if not code_url:
        raise ValueError("❌ Could not find codes/descriptions file.")
    code_path = os.path.join(base_path, os.path.basename(code_url))
    with open(code_path, 'wb') as f:
        f.write(requests.get(code_url).content)
    print(f"✅ Downloaded SEDS Codes and Descriptions")

# ==============================
# Download Total Energy Data Set
# ==============================
def download_total_energy_files():
    base_url = 'https://www.eia.gov/totalenergy/data/monthly/index.php'
    soup = BeautifulSoup(requests.get(base_url).text, 'html.parser')

    # Extract the release year from page text
    match = re.search(r'Release Date:\s+\w+\s+\d{1,2},\s+(\d{4})', soup.get_text())
    if not match:
        raise ValueError("❌ Could not find release year.")
    year = match.group(1)
    print(f"Latest Total Energy release year: {year}")
    print(f'📂 Saving to Energy Data/ATB/{year}')

    base_path = os.path.join('Energy Data', 'Total Energy', year)
    os.makedirs(base_path, exist_ok=True)

    # Download ZIP directly to memory
    zip_tag = soup.find('a', string=re.compile(r'Download all tables ZIP', re.IGNORECASE))
    if not zip_tag or not zip_tag.get('href'):
        raise ValueError("❌ ZIP file link not found.")
    zip_url = urljoin(base_url, zip_tag['href'])

    try:
        zip_response = requests.get(zip_url)
        zip_response.raise_for_status()
        zip_bytes = io.BytesIO(zip_response.content)

        with zipfile.ZipFile(zip_bytes) as zip_ref:
            for member in zip_ref.infolist():
                if member.filename.endswith('/') or not member.filename.lower().endswith('.xlsx'):
                    continue  # Skip folders and non-Excel files

                # Read Excel file from ZIP directly
                xlsx_data = zip_ref.read(member)
                xlsx_buffer = io.BytesIO(xlsx_data)
                xlsx_name = os.path.splitext(os.path.basename(member.filename))[0]

                try:
                    wb = openpyxl.load_workbook(xlsx_buffer, data_only=True)
                    sheetnames = wb.sheetnames
                
                    if "Annual Data" in sheetnames:
                        # Only process "Annual Data" sheet
                        sheet = wb["Annual Data"]
                
                        # Read from the sheet starting from the correct position
                        label = sheet['A7'].value or 'unknown'
                        label_clean = re.sub(r'[^\w\- ]+', '', str(label)).replace(' ', '_').lower()
                
                        csv_filename = f"{label_clean}.csv"  # ❗ no prefix like "annual_data" anymore
                        # If it starts with "table_", remove "table_xxx_" prefix
                        if csv_filename.startswith('table_'):
                            # Find the first underscore after "table_", then strip up to there
                            parts = csv_filename.split('_', 2)  # Split into at most 3 parts: "table", "xxx", "rest"
                            if len(parts) == 3:
                                csv_filename = parts[2]  # Keep only "rest"
                        
                        # Prepend "total_energy_" no matter what
                        csv_filename = f"total_energy_{csv_filename}"
                        csv_path = os.path.join(base_path, csv_filename)
                
                        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
                            writer = csv.writer(f)
                            for row in sheet.iter_rows(values_only=True):
                                writer.writerow(row)
                
                        print(f"✅ Downloaded Total Energy Data: {csv_filename}")
                
                    else:
                        # If no "Annual Data" sheet, process all sheets individually (ignoring monthly)
                        for sheet_name in sheetnames:
                            if "monthly" in sheet_name.lower():
                                continue  # ❗ skip any monthly data
                
                            sheet = wb[sheet_name]
                
                            # Clean the sheet name for the filename
                            label_clean = re.sub(r'[^\w\- ]+', '', sheet_name).replace(' ', '_').lower()
                
                            csv_filename = f"{label_clean}.csv"  # ❗ clean filename
                            # If it starts with "table_", remove "table_xxx_" prefix
                            if csv_filename.startswith('table_'):
                                # Find the first underscore after "table_", then strip up to there
                                parts = csv_filename.split('_', 2)  # Split into at most 3 parts: "table", "xxx", "rest"
                                if len(parts) == 3:
                                    csv_filename = parts[2]  # Keep only "rest"
                            
                            # Prepend "total_energy_" no matter what
                            csv_filename = f"total_energy_{csv_filename}"
                            csv_path = os.path.join(base_path, csv_filename)
                
                            with open(csv_path, 'w', newline='', encoding='utf-8') as f:
                                writer = csv.writer(f)
                                for row in sheet.iter_rows(values_only=True):
                                    writer.writerow(row)
                
                            print(f"✅ Downloaded Total Energy Data: {csv_filename}")
                
                except Exception as e:
                    print(f"❌ Failed to process {member.filename}: {e}")
                
    except requests.RequestException as e:
        print(f"❌ Failed to download ZIP archive: {e}")
        return

    # Download Glossary PDF
    pdf_tag = soup.find('a', href=re.compile(r'PDF', re.IGNORECASE), attrs={'title': 'Glossary'})
    if not pdf_tag or not pdf_tag.get('href'):
        raise ValueError("❌ Glossary PDF not found.")
    pdf_url = urljoin(base_url, pdf_tag['href'])
    pdf_path = os.path.join(base_path, os.path.basename(pdf_url))

    try:
        with open(pdf_path, 'wb') as f:
            f.write(requests.get(pdf_url).content)
        print(f"✅ Downloaded Total Energy Glossary PDF")
    except Exception as e:
        print(f"❌ Failed to download Glossary PDF: {e}")

# ===========================================
# Download Latest ATB Workbook & Documentation
# ===========================================
def download_atb_files():
    base_path_root = os.path.join('Energy Data', 'ATB')
    base_url_template = 'https://atb.nrel.gov/electricity/{year}/data'
    doc_url = 'https://raw.githubusercontent.com/openEDI/documentation/main/ATB.md'

    latest_year = None
    for year in range(2025, 2014, -1):
        url = base_url_template.format(year=year)
        try:
            if requests.head(url, allow_redirects=True, timeout=5).status_code == 200:
                latest_year = str(year)
                break
        except requests.RequestException:
            continue
    if not latest_year:
        raise ValueError("❌ No valid ATB year found.")
    print(f"Latest ATB year: {latest_year}")
    print(f'📂 Saving to Energy Data/ATB/{latest_year}')

    base_path = os.path.join(base_path_root, latest_year)
    os.makedirs(base_path, exist_ok=True)

    soup = BeautifulSoup(requests.get(base_url_template.format(year=latest_year)).text, 'html.parser')
    csv_tag = soup.find('a', string=re.compile(rf'Download the {latest_year} ATB Summary CSV Files', re.IGNORECASE))
    if not csv_tag or not csv_tag.get('href'):
        raise ValueError("❌ ATB Summary CSV not found.")
    csv_url = urljoin(base_url_template.format(year=latest_year), csv_tag['href'])
    csv_path = os.path.join(base_path, os.path.basename(csv_url))
    with open(csv_path, 'wb') as f:
        f.write(requests.get(csv_url).content)
    print(f"✅ Downloaded ATB Summary CSV")

    doc_path = os.path.join(base_path, 'ATB.md')
    with open(doc_path, 'wb') as f:
        f.write(requests.get(doc_url).content)
    print(f"✅ Downloaded ATB Documentation")

# ============================
# Download RMI Dataset Files
# ============================
def download_rmi_files():
    base_url = 'https://utilitytransitionhub.rmi.org/data-download/'
    base_dir = os.path.join('Energy Data', 'RMI')
    os.makedirs(base_dir, exist_ok=True)

    # Files to keep (original base names)
    keep_basenames = [
        'employees',
        'operations_emissions_by_fuel',
        'revenue_by_tech',
        'utility_state_map'
    ]

    try:
        soup = BeautifulSoup(requests.get(base_url).text, 'html.parser')
    except requests.RequestException as e:
        print(f"❌ Failed to retrieve RMI page: {e}")
        return

    print('📂 Saving to Energy Data/RMI')

    containers = soup.find_all('div', class_='container mb-16')[0]
    for container in containers:
        # Extract "Last updated" year
        match = re.search(r'Last updated:\s+\w+\s+\d{1,2},\s+(\d{4})', container.get_text())
        modified_year = match.group(1) if match else 'unknown'

        for a in container.find_all('a', href=True):
            file_url = urljoin(base_url, a['href'])
            file_name = os.path.basename(file_url)

            # Only process if .csv, .xlsx, or .zip
            if not re.search(r'\.(csv|xlsx|zip)$', file_name, re.IGNORECASE):
                continue

            base, ext = os.path.splitext(file_name)
            base_lower = base.lower()

            # Skip if not in keep list
            if not any(target in base_lower for target in keep_basenames):
                continue

            try:
                response = requests.get(file_url)
                response.raise_for_status()

                # Handle zip extraction in memory
                if ext.lower() == '.zip':
                    zip_bytes = io.BytesIO(response.content)
                    try:
                        with zipfile.ZipFile(zip_bytes, 'r') as zip_ref:
                            for member in zip_ref.infolist():
                                if member.filename.endswith('/'):
                                    continue  # Skip folders
                                inner_name = os.path.basename(member.filename)
                                inner_base, inner_ext = os.path.splitext(inner_name)

                                # Check again inside zip
                                inner_base_lower = inner_base.lower()
                                if not any(target in inner_base_lower for target in keep_basenames):
                                    continue

                                # Add year if not already present
                                if not re.search(r'\d{4}$', inner_base):
                                    inner_name = f"{inner_base}_{modified_year}{inner_ext}"

                                file_path = os.path.join(base_dir, inner_name)
                                with zip_ref.open(member) as source, open(file_path, 'wb') as target:
                                    target.write(source.read())
                                print(f"✅ Downloaded RMI Data: {inner_name}")
                    except zipfile.BadZipFile:
                        print(f"❌ Invalid ZIP: {file_name}")
                else:
                    # Non-zip file
                    if not re.search(r'\d{4}$', base):
                        file_name = f"{base}_{modified_year}{ext}"
                    file_path = os.path.join(base_dir, file_name)
                    with open(file_path, 'wb') as f:
                        f.write(response.content)
                    print(f"✅ Downloaded RMI Data: {file_name}")

            except requests.RequestException as e:
                print(f"❌ Failed to download RMI Data: {file_url}: {e}")

def main():
    try:
        print('Downloading Energy Data ...')
        print('--------------------------------------------------------')
        download_recs_files()
        print('--------------------------------------------------------')
        download_seds_files()
        print('--------------------------------------------------------')
        download_total_energy_files()
        print('--------------------------------------------------------')
        download_atb_files()
        print('--------------------------------------------------------')
        download_rmi_files()
    except Exception as e:
        print(f'An error occurred: {e}')
# Set base folders
energy_data_folder = 'Energy Data'
states = [
    'AK', 'AL', 'AR', 'AZ', 'CA', 'CO', 'CT', 'DC', 'DE', 'FL', 'GA',
    'HI', 'IA', 'ID', 'IL', 'IN', 'KS', 'KY', 'LA', 'MA', 'MD', 'ME',
    'MI', 'MN', 'MO', 'MS', 'MT', 'NC', 'ND', 'NE', 'NH', 'NJ', 'NM',
    'NV', 'NY', 'OH', 'OK', 'OR', 'PA', 'RI', 'SC', 'SD', 'TN', 'TX',
    'UT', 'VA', 'VT', 'WA', 'WI', 'WV', 'WY'
]

# =========================
# Load ATB Data
# =========================

def load_atb():
    atb_folder = os.path.join(energy_data_folder, 'ATB')
    years = [f for f in os.listdir(atb_folder) if f.isdigit()]
    latest_year = max(years)
    atb_path = os.path.join(atb_folder, latest_year)

    csv_files = [f for f in os.listdir(atb_path) if f.endswith('.csv')]
    if not csv_files:
        raise ValueError(f"❌ No CSV files found in {atb_path}")

    atb_file = os.path.join(atb_path, csv_files[0])
    print(f"✅ Loading ATB file: {atb_file}")

    atb_df = pd.read_csv(atb_file, low_memory=False, index_col=0).reset_index(drop=True)

    # Confirm year column
    first_col = atb_df.columns[0]
    if 'year' not in first_col.lower():
        raise ValueError(f"❌ First column '{first_col}' does not contain 'year'.")

    # Step 1: Remove 'default' column (no longer needed)
    atb_df = atb_df.drop(columns=['default'])

    # Step 2: Create final metadata dict per row
    metadata_rows = []
    for _, row in atb_df.iterrows():
        year = row[first_col]
        metadata = row.drop(first_col).to_dict()

        metadata_rows.append({
            'year': int(year),
            'state': 'US',
            'metadata': {'ATB': metadata}
        })

    # Step 3: Build final DataFrame
    final_atb_df = pd.DataFrame(metadata_rows)

    print("✅ Final ATB metadata dataframe ready!")
    return final_atb_df

# =========================
# Load RECS Data
# =========================

def load_recs():
    recs_folder = os.path.join(energy_data_folder, 'RECS')
    years = [f for f in os.listdir(recs_folder) if f.isdigit()]
    latest_year = max(years)
    saved_year = int(latest_year)
    recs_path = os.path.join(recs_folder, latest_year)

    recs_files = [f for f in os.listdir(recs_path) if f.endswith('.csv') and 'recs' in f.lower()]
    version_files = []
    for f in recs_files:
        match = re.search(r'v(\d+)', f.lower())
        if match:
            version_files.append((int(match.group(1)), f))
    if not version_files:
        raise ValueError(f"❌ No RECS files with version found in {recs_path}")

    highest_version_file = max(version_files, key=lambda x: x[0])[1]
    recs_file = os.path.join(recs_path, highest_version_file)
    print(f"✅ Loading RECS file: {recs_file}")

    # Load RECS file
    recs_df = pd.read_csv(recs_file, low_memory=False)

    if 'state_postal' not in recs_df.columns:
        raise ValueError("❌ 'state_postal' column missing in RECS file")

    # Rename for clarity
    recs_df = recs_df.rename(columns={'state_postal': 'state'})

    # Step 1: Create metadata dictionaries
    metadata_rows = []
    for _, row in recs_df.iterrows():
        state = row['state']
        metadata = row.drop('state').to_dict()

        # Overwrite/add year manually to be safe
        metadata['year'] = saved_year

        metadata_rows.append({
            'year': saved_year,
            'state': state,
            'metadata': {'RECS': metadata}
        })

    # Step 2: Build final DataFrame
    final_recs_df = pd.DataFrame(metadata_rows)

    print("✅ Final RECS metadata dataframe ready!")
    return final_recs_df

# =========================
# Load RMI Operations Emissions
# =========================

def load_rmi():
    # --- Setup ---
    rmi_folder = os.path.join(energy_data_folder, 'RMI')
    print("✅ RMI folder found.")

    def make_metadata(df, keyname):
        df = df.copy()
        df['year'] = df['year'].astype(int)
        df['metadata'] = df.drop(columns=['year', 'state']).to_dict(orient='records')
        df['metadata'] = df['metadata'].apply(lambda m: {keyname: m})
        return df[['year', 'state', 'metadata']]

    def load_rmi_file(pattern, state_col=None, keyname=None, rename_state=False, is_utility_map=False):
        matches = [f for f in os.listdir(rmi_folder) if pattern in f.lower()]
        if not matches:
            print(f"⚠️ No file matching pattern '{pattern}' found.")
            return None

        path = os.path.join(rmi_folder, matches[0])
        print(f"✅ Loading RMI file: {path}")
        df = pd.read_csv(path, dtype=str)

        # Special case: utility_state_map needs custom renaming
        if is_utility_map:
            df = df.rename(columns={'state': 'state_name', 'state_abbr': 'state'})
        elif state_col:
            if rename_state:
                df = df.rename(columns={state_col: 'state'})
            else:
                df['state'] = df[state_col]
        else:
            df['state'] = "NA"

        df['year'] = df['year'].astype(int)

        # Convert all numeric columns
        numeric_cols = [col for col in df.columns if col not in ['year', 'state']]
        for col in numeric_cols:
            df[col] = pd.to_numeric(df[col], errors='coerce')

        return make_metadata(df, keyname)

    # --- Load and process operations_emissions_by_fuel ---
    operations_file = [f for f in os.listdir(rmi_folder) if 'operations_emissions_by_fuel' in f.lower()]
    if not operations_file:
        raise ValueError("❌ No RMI operations_emissions_by_fuel file found.")
    operations_path = os.path.join(rmi_folder, operations_file[0])
    print(f"✅ Loading RMI Operations file: {operations_path}")
    operations_df = pd.read_csv(operations_path, dtype=str)

    operations_df['year'] = operations_df['year'].astype(int)
    numeric_columns = [col for col in operations_df.columns if col not in ['year', 'state']]
    for col in numeric_columns:
        operations_df[col] = pd.to_numeric(operations_df[col], errors='coerce')

    operations_df_final = make_metadata(operations_df, keyname='RMI_emissions')

    # --- Load additional RMI datasets ---
    revenues_df = load_rmi_file("revenue_by_tech", keyname="RMI_revenues")
    employees_df = load_rmi_file("employees", keyname="RMI_employees")
    utility_map_df = load_rmi_file("utility_state_map", keyname="RMI_utility_state_map", is_utility_map=True)

    # --- Combine all ---
    all_parts = [operations_df_final, revenues_df, employees_df, utility_map_df]
    combined = pd.concat([df for df in all_parts if df is not None], ignore_index=True)

    print("✅ Final RMI metadata dataframe ready!")
    return combined

# =========================
# Load Total Energy Data
# =========================

def load_total_energy():
    total_folder = os.path.join(energy_data_folder, 'Total Energy/2025')
    files = [f for f in os.listdir(total_folder) if f.endswith('.csv') and f != 'total_energy_contents.csv']
    all_dfs = []
    
    for file_name in files:
        file_path = os.path.join(total_folder, file_name)
    
        try:
            # Quick scan to detect if 'YYYYMM' format appears (in 3rd row)
            preview = pd.read_csv(file_path, nrows=5, header=None, dtype=str)
            header_preview = preview.iloc[2].fillna('').tolist()
    
            first_col = header_preview[0]
    
            if 'yyyymm' in str(first_col).lower():
                # ✅ State-organized special case
    
                # Step 1: Read with header at 3rd row (skip first 2 rows)
                df = pd.read_csv(file_path, skiprows=2, dtype=str)
    
                # Step 2: Clean header
                df.columns = [col.replace(',', '').strip() if isinstance(col, str) else col for col in df.columns]
    
                # Step 3: Drop 'US' column if it exists
                df = df.drop(columns=['US'], errors='ignore')
    
                # Step 4: Reformat 'YYYYMM' to 'YYYY'
                original_year_col = df.columns[0]
                df[original_year_col] = df[original_year_col].str[:4]
                df = df.rename(columns={original_year_col: 'year'})
    
                # Step 5: Melt into long format
                value_col_name = file_name.replace('.csv', '').lower()
                df_melted = df.melt(id_vars=['year'], var_name='state', value_name=value_col_name)
    
                # Step 6: Convert values to numeric
                df_melted[value_col_name] = pd.to_numeric(df_melted[value_col_name], errors='coerce')
    
                # Step 7: Group by (year, state) and take sum
                df_grouped = df_melted.groupby(['year', 'state'], as_index=True).sum()
    
                # Step 8: Sort (optional polish)
                df_grouped = df_grouped.sort_index()
    
                # Add to the list
                all_dfs.append(df_grouped)
    
                print(f"✅ Processed file: {file_name}")
    
            else:
                # ✅ Normal file case
    
                # Step 1: Read header rows manually
                df_header1 = pd.read_csv(file_path, skiprows=8, nrows=1, header=None, dtype=str)
                df_header2 = pd.read_csv(file_path, skiprows=9, nrows=1, header=None, dtype=str)
    
                # Step 2: Combine headers
                combined_headers = []
                for h1, h2 in zip(df_header1.iloc[0], df_header2.iloc[0]):
                    h1_clean = str(h1).strip().replace(',', '').replace(' ', '_').lower()
                    h2_clean = str(h2).strip().replace(',', '').replace(' ', '_').lower()
                    if h2_clean and h2_clean.lower() != 'nan':
                        combined_headers.append(f"{h1_clean}_{h2_clean}")
                    else:
                        combined_headers.append(h1_clean)
    
                # Step 3: Read full data
                df = pd.read_csv(file_path, skiprows=10, names=combined_headers, dtype=str)
    
                # Step 4: Clean column names
                col_a_original = combined_headers[0]
    
                # Step 5: Rename first column to 'year'
                df = df.rename(columns={col_a_original: 'year'})
    
                # Step 6: Rename other columns properly
                cleaned_columns = {}
                for col in df.columns:
                    if col != 'year':
                        cleaned_columns[col] = f"{file_name.replace('.csv', '').lower()}_{col}_{col_a_original}"
                df = df.rename(columns=cleaned_columns)
    
                df['state'] = 'US'
                df = df.set_index(['year', 'state'])
    
                all_dfs.append(df)
    
                print(f"✅ Processed file: {file_name}")
    
        except Exception as e:
            print(f"❌ Failed to process {file_name}: {e}")
    
    # Merge all DataFrames efficiently
    if all_dfs:
        final_df = reduce(lambda left, right: left.join(right, how='outer'), all_dfs)
        final_df = final_df.sort_index()
    
        print("✅ Final Total Energy dataframe ready!")
    
    else:
        final_df = pd.DataFrame()
        print("❌ No data was loaded.")

    return final_df

# =========================
# Load SEDS Data
# =========================

def load_seds():
    seds_folder = os.path.join(energy_data_folder, 'SEDS')
    years = [f for f in os.listdir(seds_folder) if f.isdigit()]
    latest_year = max(years)
    seds_path = os.path.join(seds_folder, latest_year)

    csv_files = [f for f in os.listdir(seds_path) if f.endswith('.csv') and 'complete' in f.lower()]
    if not csv_files:
        raise ValueError(f"❌ No 'complete' CSV files found in {seds_path}")

    seds_file = os.path.join(seds_path, csv_files[0])
    print(f"✅ Loading SEDS file: {seds_file}")

    seds_df = pd.read_csv(seds_file, low_memory=False)

    if 'Data_Status' in seds_df.columns:
        seds_df = seds_df.drop(columns=['Data_Status'])

    seds_pivoted = seds_df.pivot_table(index=['Year', 'StateCode'], columns='MSN', values='Data').reset_index()
    seds_pivoted.columns.name = None
    seds_pivoted = seds_pivoted.rename(columns={'Year': 'year', 'StateCode': 'state'})
    final_seds_df = seds_pivoted.set_index(['year', 'state'])

    print("✅ Final SEDS dataframe ready!")
    return final_seds_df

# =========================
# Merge Total Energy & SEDS
# =========================

def merge_seds_and_total(seds_df, total_df):
    # 🛠 Step 1: Standardize both datasets
    def standardize_index(df):
        if isinstance(df.index, pd.MultiIndex):
            df = df.reset_index()
        else:
            df = df.reset_index()

        if 'year' not in df.columns or 'state' not in df.columns:
            raise ValueError("❌ DataFrame missing 'year' or 'state' for merging.")

        df['year'] = df['year'].astype(int)
        df['state'] = df['state'].astype(str)

        return df.set_index(['year', 'state'])

    seds_std = standardize_index(seds_df)
    total_std = standardize_index(total_df)

    # 🛠 Step 2: Merge SEDS and Total Energy
    merged = seds_std.join(total_std, how='outer')

    # 🛠 Step 3: Reset index (now year and state are columns)
    merged = merged.reset_index()

    # 🛠 Step 4: Build metadata column
    priority_cols = ['year', 'state']
    data_cols = [col for col in merged.columns if col not in priority_cols]

    def row_to_metadata(row):
        metadata = {col: row[col] for col in data_cols if pd.notna(row[col])}
        return {"SEDS_AND_TOTAL_ENERGY": metadata}

    merged['metadata'] = merged.apply(row_to_metadata, axis=1)

    # 🛠 Step 5: Keep only year, state, metadata
    final_df = merged[['year', 'state', 'metadata']]

    print(f"✅ Merged SEDS and Total Energy into metadata dataframe")
    return final_df
    

# ─── Arkansas ────────────────────────────────────────────────────────────────

def extract_history_marker_ark(text: str) -> Union[int, str, None]:
    u4 = re.findall(r'\u2002(\d{4})', text)
    if u4:
        return int(u4[-1])
    parts = re.split(r'History', text, maxsplit=1, flags=re.IGNORECASE)
    history = parts[1] if len(parts) > 1 else text
    m = re.search(r'Acts\s+(\d{4})', history)
    if m:
        return int(m.group(1))
    m2 = re.search(r'Source:\s*L\.\s*(\d{4})', history)
    if m2:
        return int(m2.group(1))
    u2 = re.findall(r'\u2002(\d{2})', text)
    if u2:
        return int(u2[-1])
    m3 = re.search(r'Acts\s+(\d{2})', history)
    if m3:
        return int(m3.group(1))
    m4 = re.search(r'Source:\s*L\.\s*(\d{2})', history)
    if m4:
        return int(m4.group(1))
    if re.search(r'\[Reserved\.\]', history, re.IGNORECASE):
        return "Reserved"
    if re.search(r'\[Repealed\.\]', history, re.IGNORECASE):
        return "Repealed"
    return None

def trim_to_body_ark(text: str) -> str:
    marker = "\n\n\n"
    idx = text.find(marker)
    if idx != -1:
        return text[idx+len(marker):]
    parts = re.split(r'(?:\r?\n){3,}', text, maxsplit=1)
    return parts[1] if len(parts)>1 else text

def scrape_arkansas_df() -> pd.DataFrame:
    texts: List[str] = []
    start = time.time()
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()
        page.goto(
            "https://advance.lexis.com/container?config=00JAA3ZTU0NTIzYy0zZDEyLTRhYmQtYmRmMS1iMWIxNDgxYWMxZTQK"
            "AFBvZENhdGFsb2cubRW4ifTiwi5vLw6cI1uX&crid=8efec3df-f9c3-48dd-af22-2d09db950145", 
            wait_until="networkidle"
        )
        try:
            page.wait_for_selector("input#btnagreeterms", timeout=5_000)
            page.click("input#btnagreeterms")
        except PWTimeout:
            pass

        page.fill("textarea#searchTerms", "energy")
        page.press("textarea#searchTerms","Enter")
        page.wait_for_load_state("networkidle")

        while True:
            if time.time()-start > 600:
                break
            page.wait_for_selector("li.usview", timeout=20_000)
            hits = page.locator("li.usview")
            n = hits.count()
            for i in range(n):
                if time.time()-start>600:
                    break
                hits.nth(i).locator("p.min.vis a").first.click()
                page.wait_for_load_state("networkidle")
                page.wait_for_selector("section#document", timeout=10_000)
                texts.append(page.inner_text("section#document"))
                page.go_back()
                page.wait_for_load_state("networkidle")
            time.sleep(1)
            try:
                nxt = page.locator("nav.pagination >> a:has-text('Next')")
                if nxt.count() and nxt.first.is_visible():
                    nxt.first.click()
                    page.wait_for_load_state("networkidle")
                    continue
            except PWTimeout:
                pass
            break
        browser.close()

    # dedupe + build DF
    seen = set(); uniq = []
    for t in texts:
        if t not in seen:
            seen.add(t); uniq.append(t)
    df = pd.DataFrame(uniq, columns=["text"])
    df["year"] = df["text"].apply(extract_history_marker_ark)
    df["text"] = df["text"].apply(trim_to_body_ark)
    df = df[df["year"].notnull()].copy()
    df["year"] = df["year"].apply(lambda y: (1900+y) if isinstance(y,int) and y<100 else y)
    df["location"] = "Arkansas"
    return df[["year","location","text"]]

# ─── Georgia ────────────────────────────────────────────────────────────────

def extract_history_marker_ga(text: str) -> Union[int, str, None]:
    m = re.search(r'Ga\.\s*L\.\s*(\d{4})', text)
    if m:
        return int(m.group(1))
    if re.search(r'\[Reserved\.\]', text, re.IGNORECASE):
        return "Reserved"
    if re.search(r'\[Repealed\.\]', text, re.IGNORECASE):
        return "Repealed"
    return None

def trim_to_body_ga(text: str) -> str:
    marker="\n\n\n"; idx=text.find(marker)
    if idx!=-1: return text[idx+len(marker):]
    parts=re.split(r'(?:\r?\n){3,}', text, maxsplit=1)
    return parts[1] if len(parts)>1 else text

def scrape_georgia_df() -> pd.DataFrame:
    texts=[]; start=time.time()
    with sync_playwright() as p:
        browser=p.chromium.launch(headless=True)
        page=browser.new_page()
        page.goto(
            "https://advance.lexis.com/container?config=00JAAzZDgzNzU2ZC05MDA0LTRmMDItYjkzMS0xOGY3MjE3OWNlODIK"
            "AFBvZENhdGFsb2fcIFfJnJ2IC8XZi1AYM4Ne&crid=d1ef0e4a-f560-4a3f-bca1-09d66269998b",
            wait_until="networkidle"
        )
        try:
            page.wait_for_selector("input#btnagreeterms", timeout=5_000)
            page.click("input#btnagreeterms")
        except PWTimeout:
            pass

        page.fill("textarea#searchTerms","energy")
        page.press("textarea#searchTerms","Enter")
        page.wait_for_load_state("networkidle")

        while True:
            if time.time()-start>600: break
            page.wait_for_selector("li.usview", timeout=20_000)
            hits=page.locator("li.usview"); n=hits.count()
            for i in range(n):
                if time.time()-start>600: break
                try:
                    hits.nth(i).locator("p.min.vis a").first.click()
                    page.wait_for_load_state("networkidle")
                    page.wait_for_selector("section#document",timeout=10_000)
                    texts.append(page.inner_text("section#document"))
                except:
                    texts.append("")
                finally:
                    try:
                        page.go_back(); page.wait_for_load_state("networkidle")
                    except: pass
            time.sleep(1)
            try:
                nxt=page.locator("nav.pagination >> a[data-action='nextpage']").first
                if nxt.count() and nxt.is_visible():
                    nxt.click(); page.wait_for_load_state("networkidle"); continue
            except PWTimeout:
                pass
            break
        browser.close()

    seen=set(); uniq=[]
    for t in texts:
        if t not in seen:
            seen.add(t); uniq.append(t)
    df=pd.DataFrame(uniq,columns=["text"])
    df["year"]=df["text"].apply(extract_history_marker_ga)
    df=df[df["year"].notnull()].copy()
    df["text"]=df["text"].apply(trim_to_body_ga)
    df["year"]=df["year"].apply(lambda y:(1900+y) if isinstance(y,int) and y<100 else y)
    df["location"]="Georgia"
    return df[["year","location","text"]]

# ─── Colorado ───────────────────────────────────────────────────────────────

def extract_history_marker_co(text:str)->Union[int,str,None]:
    # same logic as Arkansas
    return extract_history_marker_ark(text)

def trim_to_body_co(text:str)->str:
    return trim_to_body_ark(text)

def scrape_colorado_df()->pd.DataFrame:
    texts=[]; start=time.time()
    with sync_playwright() as p:
        browser=p.chromium.launch(headless=True)
        page=browser.new_page()
        page.goto(
            "https://advance.lexis.com/container?config=00JAA3ZTU0NTIzYy0zZDEyLTRhYmQtYmRmMS1iMWIxNDgxYWMxZTQK"
            "AFBvZENhdGFsb2cubRW4ifTiwi5vLw6cI1uX&crid=71f400f1-686d-4c50-8ecc-7711eca7c5a8",
            wait_until="networkidle"
        )
        try:
            page.wait_for_selector("input#btnagreeterms",timeout=5_000)
            page.click("input#btnagreeterms")
        except PWTimeout:
            pass

        page.fill("textarea#searchTerms","energy")
        page.press("textarea#searchTerms","Enter")
        page.wait_for_load_state("networkidle")

        while True:
            if time.time()-start>600: break
            page.wait_for_selector("li.usview",timeout=20_000)
            hits=page.locator("li.usview"); n=hits.count()
            for i in range(n):
                if time.time()-start>600: break
                hits.nth(i).locator("p.min.vis a").first.click()
                page.wait_for_load_state("networkidle")
                page.wait_for_selector("section#document",timeout=10_000)
                texts.append(page.inner_text("section#document"))
                page.go_back(); page.wait_for_load_state("networkidle")
            time.sleep(1)
            try:
                nxt=page.locator("nav.pagination >> a:has-text('Next')")
                if nxt.count() and nxt.first.is_visible():
                    nxt.first.click(); page.wait_for_load_state("networkidle"); continue
            except PWTimeout:
                pass
            break
        browser.close()

    seen=set(); uniq=[]
    for t in texts:
        if t not in seen:
            seen.add(t); uniq.append(t)
    df=pd.DataFrame(uniq,columns=["text"])
    df["year"]=df["text"].apply(extract_history_marker_co)
    df["text"]=df["text"].apply(trim_to_body_co)
    df=df[df["year"].notnull()].copy()
    df["year"]=df["year"].apply(lambda y:(1900+y) if isinstance(y,int) and y<100 else y)
    df["location"]="Colorado"
    return df[["year","location","text"]]

# ─── Delaware ───────────────────────────────────────────────────────────────

def extract_last_approval_year_de(text:str)->Optional[int]:
    pattern = r"""(?:Approved|Passed(?:\s+at\s+[^,]+)?),?\s*[A-Za-z]+\s+\d{1,2}(?:[.,]\s*)(?:A\.\s*D\.\s*)?(\d{4})"""
    m = list(re.finditer(pattern,text,flags=re.IGNORECASE|re.VERBOSE))
    return int(m[-1].group(1)) if m else None

def scrape_delaware_df()->pd.DataFrame:
    # step 1: gather section links
    section_urls=[]; start=time.time()
    with sync_playwright() as p:
        b=p.chromium.launch(headless=True); pg=b.new_page(viewport={"width":1280,"height":800})
        pg.goto("https://delcode.delaware.gov/",wait_until="networkidle")
        pg.fill("input#srch-term","energy"); pg.click("button[onclick='submitSearch()']")
        pg.wait_for_selector("#SearchResultsGrid tbody tr.k-master-row",timeout=15_000)
        page_num=1
        while True:
            if time.time()-start>600: break
            rows=pg.locator("#SearchResultsGrid tbody tr.k-master-row"); cnt=rows.count()
            for i in range(cnt):
                if time.time()-start>600: break
                cells=rows.nth(i).locator("td[role='gridcell']")
                if cells.count()>=4:
                    link=cells.nth(3).locator("a").first
                    href=link.get_attribute("href"); txt=link.inner_text().strip()
                    if href:
                        full=href if href.startswith("http") else "https://delcode.delaware.gov"+href
                        section_urls.append((full,txt))
            # next?
            if pg.locator("li.k-pager-nav.k-state-disabled").count(): break
            pg.click("a.k-link.k-pager-nav[title='Go to the next page']")
            pg.wait_for_selector("#SearchResultsGrid tbody tr.k-master-row",timeout=15_000)
            page_num+=1
        b.close()

    # step 2: get chapter links per section
    chap_urls=[]; start=time.time()
    for url,sec in section_urls:
        if time.time()-start>600: break
        with sync_playwright() as p:
            b=p.chromium.launch(headless=True); pg=b.new_page()
            pg.goto(url,wait_until="domcontentloaded")
            js=f"""() => {{
              const head=document.querySelector("div.SectionHead[id='{sec}']");
              if(!head)return[];
              const s=head.closest("div.Section");
              return Array.from(s.querySelectorAll("a")).map(a=>[a.href,a.textContent.trim()]);
            }}"""
            raw=pg.evaluate(js); b.close()
        for href, _ in raw:
            if href not in chap_urls:
                chap_urls.append(href)

    # step 3: fetch body + extract year
    docs=[]; years=[]
    start=time.time()
    for u in chap_urls:
        if time.time()-start>600: break
        with sync_playwright() as p:
            b=p.chromium.launch(headless=True); pg=b.new_page()
            pg.goto(u,wait_until="domcontentloaded")
            pg.wait_for_selector("div#chapterBody",timeout=10_000)
            txt=pg.inner_text("div#chapterBody"); b.close()
        docs.append(txt); years.append(extract_last_approval_year_de(txt))

    df=pd.DataFrame({"year":years,"text":docs}).dropna(subset=["year"])
    df["location"]="Delaware"
    return df[["year","location","text"]]

# ─── Florida ────────────────────────────────────────────────────────────────

def extract_original_enactment_year_fl(h:str)->Optional[int]:
    parts=re.split(r'\n\n—\s*History\s*—\n\n',h,flags=re.IGNORECASE)
    if len(parts)<2: return None
    m=re.search(r'ch\.\s*(\d{2,4})-',parts[1])
    if not m: return None
    raw=int(m.group(1))
    return (1900+raw) if raw<100 and raw>=50 else (2000+raw) if raw<50 else raw

def scrape_florida_df()->pd.DataFrame:
    links=[]; start=time.time()
    with sync_playwright() as p:
        b=p.firefox.launch(headless=True); pg=b.new_page()
        pg.goto("https://www.flsenate.gov/laws/statutes",wait_until="domcontentloaded")
        pg.fill("input#filteredData_StatuteSearchQuery","energy")
        pg.click("input[name='StatutesGoSubmit']")
        pg.wait_for_selector("table.tbl.width100 a",timeout=10_000)
        for a in pg.query_selector_all("table.tbl.width100 a"):
            href=a.get_attribute("href"); txt=a.inner_text().strip()
            if href:
                links.append((txt,f"https://www.flsenate.gov{href}"))
        b.close()

    docs=[]; years=[]
    for i,(_,url) in enumerate(links,1):
        if time.time()-start>600: break
        try:
            with sync_playwright() as p:
                b=p.chromium.launch(headless=True); pg=b.new_page()
                pg.goto(url,wait_until="domcontentloaded")
                pg.wait_for_selector("span.SectionBody div.Paragraph",timeout=10_000)
                paras=pg.locator("span.SectionBody div.Paragraph").all_text_contents()
                body="\n\n".join(p.strip() for p in paras if p.strip())
                hl=pg.locator("div.History span.HistoryText")
                hist=hl.count() and hl.inner_text().strip() or ""
                full=(f"{body}\n\n— History —\n\n{hist}" if hist else body)
                b.close()
            docs.append(full)
            years.append(extract_original_enactment_year_fl(full))
        except:
            docs.append(""); years.append(None)

    df=pd.DataFrame({"year":years,"text":docs}).dropna(subset=["year"])
    df["location"]="Florida"
    return df[["year","location","text"]]

# ─── Combine & Save state data ────────────────────────────────────────────────────────
dfs = [
    scrape_arkansas_df(),
    scrape_georgia_df(),
    scrape_colorado_df(),
    scrape_delaware_df(),
    scrape_florida_df(),
]
combined_state = pd.concat(dfs, ignore_index=True)
combined_state["metadata"]  = combined_state["text"]
combined_state["state"]  = combined_state["location"]
combined_state = combined_state[['year', 'state', 'metadata']]

def final_concat():
    # 🛠 Step 1: Concatenate all datasets
    combined_df = pd.concat([
        atb_df,
        recs_df,
        rmi_df,
        seds_and_total_df,
        combined_state
    ], axis=0, ignore_index=True)
    
    # Step 2: State abbreviation → full state name
    # Reference: https://www.usps.com/send/official-abbreviations.htm
    state_abbr_to_name = {
        'AL': 'Alabama', 'AK': 'Alaska', 'AZ': 'Arizona', 'AR': 'Arkansas',
        'CA': 'California', 'CO': 'Colorado', 'CT': 'Connecticut', 'DE': 'Delaware',
        'FL': 'Florida', 'GA': 'Georgia', 'HI': 'Hawaii', 'ID': 'Idaho',
        'IL': 'Illinois', 'IN': 'Indiana', 'IA': 'Iowa', 'KS': 'Kansas',
        'KY': 'Kentucky', 'LA': 'Louisiana', 'ME': 'Maine', 'MD': 'Maryland',
        'MA': 'Massachusetts', 'MI': 'Michigan', 'MN': 'Minnesota', 'MS': 'Mississippi',
        'MO': 'Missouri', 'MT': 'Montana', 'NE': 'Nebraska', 'NV': 'Nevada',
        'NH': 'New Hampshire', 'NJ': 'New Jersey', 'NM': 'New Mexico', 'NY': 'New York',
        'NC': 'North Carolina', 'ND': 'North Dakota', 'OH': 'Ohio', 'OK': 'Oklahoma',
        'OR': 'Oregon', 'PA': 'Pennsylvania', 'RI': 'Rhode Island', 'SC': 'South Carolina',
        'SD': 'South Dakota', 'TN': 'Tennessee', 'TX': 'Texas', 'UT': 'Utah',
        'VT': 'Vermont', 'VA': 'Virginia', 'WA': 'Washington', 'WV': 'West Virginia',
        'WI': 'Wisconsin', 'WY': 'Wyoming', 'DC': 'District of Columbia'
    }
    
    # Only replace if it matches a valid abbreviation
    combined_df['state'] = combined_df['state'].apply(
        lambda x: state_abbr_to_name.get(x, x)  # If x is a valid abbreviation, replace; else keep as is
    )
    
    # 🧹 Step 3: Optional sort
    combined_df = combined_df.sort_values(['year', 'state']).reset_index(drop=True)
    
    # ✅ Done
    print(f"✅ Final combined dataset shape: {combined_df.shape}")
    return combined_df
    
main()

atb_df = load_atb()
recs_df = load_recs()
rmi_df = load_rmi()

seds_df = load_seds()
total_df = load_total_energy()
seds_and_total_df = merge_seds_and_total(seds_df, total_df)

final_energy_df = final_concat()

return final_energy_df

law_df = law_data()
comm_df = community_data()
energy_df = energy_data()


law_df = law_df.rename(columns={'Combined': 'metadata'})

year_column_df1 = law_df.pop('year')
law_df.insert(0, 'year', year_column_df1)

comm_df = comm_df.rename(columns={'rest_of_data': 'metadata'})

final_merged = pd.merge(law_data, comm_df, on = ['year', 'state', 'metadata'], how = "outer", suffixes = ('_taxincentive', '_energydf'))
final_merged = pd.merge(final_merged, energy_df, on = ['year', 'state', 'metadata'], how = "outer", suffixes = ('', '_energydf2'))

final_merged['year'] = pd.to_numeric(final_merged['year'], errors='coerce').astype('Int64')
print(final_merged.head())
final_merged.to_csv("final_merged_df.csv", index=False) 
